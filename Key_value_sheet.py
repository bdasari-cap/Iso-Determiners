import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

def run_key_value_log_app():
    # st.set_page_config(page_title="Chemical Risk App", layout="wide")
    st.title("🧪 Object Log Management")


    EXCEL_PATH = r"C:\Users\Uday\Downloads\Key_Value Data Collector and Tracer v2.xlsx"

    # Load and normalize data
    demo_df = pd.read_excel(EXCEL_PATH, sheet_name="Demo_Object_Log")
    demo_df.columns = demo_df.columns.str.strip().str.replace(' ', '_')
    demo_df['Object_ID'] = pd.to_numeric(demo_df['Object_ID'], errors='coerce').astype('Int64')
    demo_df['Date'] = pd.to_datetime(demo_df['Date'], errors='coerce')

    library_df = pd.read_excel(EXCEL_PATH, sheet_name="Library_Log")
    library_df.columns = library_df.columns.str.strip().str.replace(' ', '_')
    library_df['Library_ID'] = pd.to_numeric(library_df['Library_ID'], errors='coerce').astype('Int64')
    library_df = library_df.dropna(subset=['Library_ID', 'Library_Name'])

    data_source_df = pd.read_excel(EXCEL_PATH, sheet_name="Data _Source_Log")
    data_source_df.columns = data_source_df.columns.str.strip().str.replace(' ', '_')
    data_source_df['Data_Source_ID'] = pd.to_numeric(data_source_df['Data_Source_ID'], errors='coerce').astype('Int64')
    data_source_df = data_source_df.dropna(subset=['Data_Source_ID', 'Data_Source_Name'])

    # Auto-increment Object ID
    existing_ids = demo_df['Object_ID'].dropna().astype(int).tolist()
    next_object_id = max(existing_ids) + 1 if existing_ids else 1



    # Main view mode options
    view_mode = st.radio("Choose an option:", ["Search Existing Records", "View All Records", "Add New Entry"])

    if view_mode == "View All Records":
        st.subheader("📋 All Records in Object_Log")
        st.dataframe(demo_df)

    elif view_mode == "Search Existing Records":
        st.subheader("Search in Object_Log")
        search_term = st.text_input("Enter keyword to search across all columns:")
        process_filter = st.selectbox("Filter by Process Status (optional):", ["All"] + sorted(demo_df['Process_Status'].dropna().unique().tolist()))

        filtered_df = demo_df.copy()

        if process_filter != "All":
            filtered_df = filtered_df[filtered_df['Process_Status'] == process_filter]

        if search_term:
            mask = filtered_df.apply(lambda row: row.astype(str).str.contains(search_term, case=False).any(), axis=1)
            filtered = filtered_df[mask]
        else:
            filtered = filtered_df

        st.write(f"### Found {len(filtered)} matching record(s):")
        st.dataframe(filtered)

        if not filtered.empty:
            selected_index = st.selectbox("Select a record to edit (by index):", filtered.index.tolist())

            row = filtered.loc[selected_index]
            st.write("### Selected Record:")
            st.write(row.to_frame().T)

            with st.form("edit_form"):
                st.markdown(f"**Anchor Object Name:** {row['Anchor_Object_Name']}")
                source_id = st.text_input("Source Local ID", value=row['Source_Local_ID'])
                object_type_subtype = st.text_input("Object Type SubType", value=row['Object_Type_SubType'])
                process_status = st.text_input("Process Status", value=row['Process_Status'])
                other_details = st.text_input("Other Details", value=row['Other_Details'])
                filename = st.text_input("Filename", value=row['Filename'])
                reference_url = st.text_input("Reference URL", value=row['Reference_URL'])

                date_value = row['Date'] if pd.notnull(row['Date']) else datetime.today()
                date = st.date_input("Date", value=date_value)

                current_lib_name = str(row['Library_Name']) if pd.notna(row['Library_Name']) else ""
                current_ds_title = str(row['Data_Source_Title']) if pd.notna(row['Data_Source_Title']) else ""

                lib_options = [""] + sorted(library_df['Library_Name'].dropna().unique().tolist())
                ds_options = [""] + sorted(data_source_df['Data_Source_Name'].dropna().unique().tolist())

                lib_index = lib_options.index(current_lib_name) if current_lib_name in lib_options else 0
                ds_index = ds_options.index(current_ds_title) if current_ds_title in ds_options else 0

                lib_name = st.selectbox("Library Name", lib_options, index=lib_index)
                ds_title = st.selectbox("Data Source Title", ds_options, index=ds_index)

                submitted_edit = st.form_submit_button("✅ Save Changes")
                submitted_delete = st.form_submit_button("🗑️ Delete Record")

            if submitted_edit:
                demo_df.at[selected_index, 'Source_Local_ID'] = source_id
                demo_df.at[selected_index, 'Object_Type_SubType'] = object_type_subtype
                demo_df.at[selected_index, 'Process_Status'] = process_status
                demo_df.at[selected_index, 'Other_Details'] = other_details
                demo_df.at[selected_index, 'Filename'] = filename
                demo_df.at[selected_index, 'Reference_URL'] = reference_url
                demo_df.at[selected_index, 'Date'] = date
                demo_df.at[selected_index, 'Library_Name'] = lib_name
                demo_df.at[selected_index, 'Data_Source_Title'] = ds_title

                try:
                    book = load_workbook(EXCEL_PATH)
                    if "Demo_Object_Log" in book.sheetnames:
                        std = book["Demo_Object_Log"]
                        book.remove(std)
                    writer = pd.ExcelWriter(EXCEL_PATH, engine='openpyxl')
                    writer.book = book
                    demo_df.to_excel(writer, sheet_name="Demo_Object_Log", index=False)
                    writer.close()
                    st.success("✅ Changes saved successfully!")
                except Exception as e:
                    st.error(f"❌ Failed to save changes: {e}")

            elif submitted_delete:
                demo_df = demo_df.drop(index=selected_index)
                try:
                    book = load_workbook(EXCEL_PATH)
                    if "Demo_Object_Log" in book.sheetnames:
                        std = book["Demo_Object_Log"]
                        book.remove(std)
                    writer = pd.ExcelWriter(EXCEL_PATH, engine='openpyxl')
                    writer.book = book
                    demo_df.to_excel(writer, sheet_name="Demo_Object_Log", index=False)
                    writer.close()
                    st.success("🗑️ Record deleted successfully!")
                except Exception as e:
                    st.error(f"❌ Failed to delete record: {e}")

    else:
        st.title("➕ Add Entry to Object_Log")

        submitted = False

        with st.form("entry_form"):
            anchor_name = st.text_input("Anchor Object Name")
            source_id = st.text_input("Source Local ID")

            file_type = st.text_input("File Type")
            doc_type = st.selectbox("Document Type", ["Company", "Sanction", "Reference", "Other"])
            object_type_subtype = f"{file_type}:{doc_type}" if file_type else ""

            process_status = st.selectbox("Process Status", [
                "Unprocessed", "Processed Fully",
                "Processed Partially (NLP/NER)", "Processed Partially (OCR)",
                "Processed Partially (Object Recognition)",
                "Processed Partially (Substance Reference Matched)", "Archived"
            ])

            other_details = st.text_input("Other Details")
            filename = st.text_input("Filename")
            reference_url = st.text_input("Reference URL")
            date = st.date_input("Date", value=datetime.today())

            derived_id_str = st.selectbox("Derived From Object ID (optional)", [""] + demo_df['Object_ID'].dropna().astype(int).astype(str).tolist(), index=0)
            derived_id = int(derived_id_str) if derived_id_str.isdigit() else ""

            derived_title, derived_type = "", ""
            if derived_id:
                match = demo_df[demo_df['Object_ID'] == derived_id]
                if not match.empty:
                    derived_title = match.iloc[0].get('Anchor_Object_Name', '')
                    derived_type = match.iloc[0].get('Object_Type_SubType', '')

            lib_name = st.selectbox("Select or Search Library Name", [""] + sorted(library_df['Library_Name'].dropna().unique().tolist()), index=0)
            lib_id = library_df[library_df['Library_Name'] == lib_name]['Library_ID'].values[0] if lib_name else ""

            ds_title = st.selectbox("Select or Search Data Source Title", [""] + sorted(data_source_df['Data_Source_Name'].dropna().unique().tolist()), index=0)
            ds_id = data_source_df[data_source_df['Data_Source_Name'] == ds_title]['Data_Source_ID'].values[0] if ds_title else ""

            st.markdown(f"📁 **Library ID**: `{lib_id}`")
            st.markdown(f"🧬 **Data Source ID**: `{ds_id}`")
            if derived_id != "":
                st.markdown(f"🧹 **Derived From Object Title**: `{derived_title if derived_title else 'Not found'}`")
                st.markdown(f"🔖 **Derived From Object Type-SubType**: `{derived_type if derived_type else 'Not found'}`")

            submitted = st.form_submit_button("✅ Save Entry")

        if submitted:
            new_row = [
                next_object_id, anchor_name or "", source_id or "", object_type_subtype or "", process_status or "",
                other_details or "", filename or "", reference_url or "", date.strftime("%d-%m-%Y") if date else "",
                lib_id, lib_name, derived_id if derived_id else "", derived_title, derived_type, ds_id, ds_title
            ]

            try:
                wb = load_workbook(EXCEL_PATH)
                ws = wb["Demo_Object_Log"]
                ws.append(new_row)
                wb.save(EXCEL_PATH)
                st.success(f"✅ Entry saved successfully with Object_ID {next_object_id}!")
            except Exception as e:
                st.error(f"❌ Failed to save entry: {e}")
